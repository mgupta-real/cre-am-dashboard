"""
services/rent_roll_parser.py
Parses standardized rent-roll Excel files.

Canonical columns (from Phoenix Commons template):
  Unit No | Unit Size (SF) | Market Rent (Monthly) | Effective Rent (Monthly)
  Move In Date | Lease Start Date | Lease End Date | Move Out Date
  Tenant Name | Unit Type
"""
from datetime import datetime, date
import pandas as pd
import numpy as np
from dateutil import parser as dparser
from openpyxl import load_workbook


# ── Status detection ─────────────────────────────────────────────────────────
VACANT_NAMES = {"vacant", "vacancy", "vacant-up", "vacant-dn", "down", "offline"}
NOTICE_NAMES = {"notice", "ntv", "notice to vacate", "giving notice"}
MODEL_NAMES  = {"model", "admin", "office", "leasing office"}


def _detect_status(tenant: str, eff_rent, move_out) -> str:
    if not tenant:
        return "Vacant"
    t = str(tenant).lower().strip()
    if any(k in t for k in VACANT_NAMES):
        return "Vacant"
    if any(k in t for k in NOTICE_NAMES):
        return "Notice"
    if any(k in t for k in MODEL_NAMES):
        return "Model/Admin"
    # Has move-out date in the future → still occupied but notice-ish
    if move_out:
        try:
            mo = dparser.parse(str(move_out)) if isinstance(move_out, str) else move_out
            if mo and mo.date() > date.today():
                return "Notice"
        except Exception:
            pass
    if eff_rent is None or (isinstance(eff_rent, (int, float)) and eff_rent == 0):
        return "Vacant"
    return "Occupied"


def _parse_date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, (datetime,)):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return dparser.parse(str(v)).date()
    except Exception:
        return None


def _safe_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        f = float(v)
        return f if not np.isnan(f) else None
    except Exception:
        return None


# ── Expiry bucketing ─────────────────────────────────────────────────────────
def _expiry_bucket(lease_end: date | None) -> str:
    if lease_end is None:
        return "Unknown"
    today = date.today()
    delta = (lease_end - today).days
    if delta < 0:
        return "Expired"
    if delta <= 90:
        return "0–3 Months"
    if delta <= 180:
        return "3–6 Months"
    if delta <= 365:
        return "6–12 Months"
    if delta <= 730:
        return "1–2 Years"
    if delta <= 1095:
        return "2–3 Years"
    return "3+ Years"


def parse_rent_roll(file_path: str) -> dict:
    """
    Parse a standardized Rent Roll Excel file.

    Returns
    -------
    {
        "as_of_date": date | None,
        "units": [
            {
                "unit_no": str,
                "unit_type": str,
                "unit_size_sf": float,
                "status": str,
                "tenant_name": str,
                "market_rent": float,
                "effective_rent": float,
                "move_in_date": date,
                "lease_start": date,
                "lease_end": date,
                "move_out_date": date,
                "rent_per_sf": float,
                "delta_amt": float,
                "delta_pct": float,
                "expiry_bucket": str,
            }, ...
        ],
        "summary": { ... },
        "unit_mix": { unit_type: { count, pct, avg_inplace, avg_market, avg_sf } },
        "lease_expirations": { "YYYY-MM": count },
        "errors": [],
        "warnings": [],
    }
    """
    errors = []
    warnings = []

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        return {"errors": [f"Cannot open file: {e}"], "warnings": []}

    # Find the right sheet
    sheet_name = None
    for sn in wb.sheetnames:
        if "standardized" in sn.lower() or "rent roll" in sn.lower():
            sheet_name = sn
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]
        warnings.append(f"Using first sheet: {sheet_name}")

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        errors.append("Sheet is empty.")
        return {"errors": errors, "warnings": warnings}

    # ── Find header row ────────────────────────────────────────────────────
    col_map = {}
    header_idx = None
    for i, row in enumerate(all_rows[:10]):
        for j, cell in enumerate(row):
            if isinstance(cell, str):
                cl = cell.lower().strip()
                if "unit no" in cl or "unit number" in cl:
                    col_map["unit_no"] = j
                    header_idx = i
                elif "unit size" in cl or "sq ft" in cl or "sqft" in cl:
                    col_map["unit_size_sf"] = j
                elif "market rent" in cl:
                    col_map["market_rent"] = j
                elif "effective rent" in cl:
                    col_map["effective_rent"] = j
                elif "move in" in cl or "move-in" in cl:
                    col_map["move_in_date"] = j
                elif "lease start" in cl:
                    col_map["lease_start"] = j
                elif "lease end" in cl or "expir" in cl:
                    col_map["lease_end"] = j
                elif "move out" in cl or "move-out" in cl:
                    col_map["move_out_date"] = j
                elif "tenant" in cl or "resident" in cl:
                    col_map["tenant_name"] = j
                elif "unit type" in cl or "floorplan" in cl or "type" in cl:
                    if "unit_type" not in col_map:
                        col_map["unit_type"] = j
        if header_idx is not None:
            break

    if header_idx is None:
        errors.append("Could not locate header row in rent roll.")
        return {"errors": errors, "warnings": warnings}

    required = ["unit_no", "market_rent", "effective_rent"]
    missing = [k for k in required if k not in col_map]
    if missing:
        errors.append(f"Missing required columns: {missing}")
        return {"errors": errors, "warnings": warnings}

    # ── Parse unit rows ───────────────────────────────────────────────────
    units = []
    STOP_WORDS = {"total", "totals", "average", "averages", "notes", "subtotal"}

    for row in all_rows[header_idx + 1:]:
        if all(v is None or v == "" for v in row):
            continue
        unit_no = str(row[col_map["unit_no"]]).strip() if col_map.get("unit_no") is not None and row[col_map["unit_no"]] is not None else None
        if not unit_no or unit_no.lower() in ("none", ""):
            continue
        # Stop at summary/totals/notes rows — anything below TOTALS line is not a real unit
        if any(sw in unit_no.lower() for sw in STOP_WORDS):
            break
        # Also stop if unit_no starts with bullet or special char (notes section)
        if unit_no.startswith(("•", "*", "#", "Note")):
            break
        # Skip legend/placeholder rows that have no SF and no market rent
        # (e.g. unit-type legend rows like B04, C1 at bottom of some rent rolls)
        raw_sf     = row[col_map["unit_size_sf"]]   if "unit_size_sf" in col_map else None
        raw_market = row[col_map["market_rent"]]     if "market_rent"  in col_map else None
        if raw_sf is None and raw_market is None:
            continue

        sf           = _safe_float(row[col_map["unit_size_sf"]]) if "unit_size_sf" in col_map else None
        market_rent  = _safe_float(row[col_map["market_rent"]])
        eff_rent     = _safe_float(row[col_map["effective_rent"]])
        tenant       = str(row[col_map["tenant_name"]]).strip() if "tenant_name" in col_map and row[col_map["tenant_name"]] is not None else ""
        unit_type    = str(row[col_map["unit_type"]]).strip()   if "unit_type"    in col_map and row[col_map["unit_type"]]    is not None else "Unknown"
        move_in      = _parse_date(row[col_map["move_in_date"]]) if "move_in_date"  in col_map else None
        lease_start  = _parse_date(row[col_map["lease_start"]])  if "lease_start"   in col_map else None
        lease_end    = _parse_date(row[col_map["lease_end"]])    if "lease_end"     in col_map else None
        move_out     = _parse_date(row[col_map["move_out_date"]]) if "move_out_date" in col_map else None

        status = _detect_status(tenant, eff_rent, move_out)

        # Effective rent = 0 for vacant
        if status == "Vacant":
            eff_rent = 0.0

        # Computed fields
        rent_per_sf = (eff_rent / sf) if (eff_rent and sf and sf > 0) else None
        delta_amt   = (market_rent - eff_rent) if (market_rent and eff_rent is not None) else None
        delta_pct   = (delta_amt / market_rent * 100) if (delta_amt is not None and market_rent and market_rent > 0) else None
        expiry_buck = _expiry_bucket(lease_end)

        units.append({
            "unit_no":       unit_no,
            "unit_type":     unit_type,
            "unit_size_sf":  sf,
            "status":        status,
            "tenant_name":   tenant,
            "market_rent":   market_rent,
            "effective_rent":eff_rent,
            "move_in_date":  move_in,
            "lease_start":   lease_start,
            "lease_end":     lease_end,
            "move_out_date": move_out,
            "rent_per_sf":   rent_per_sf,
            "delta_amt":     delta_amt,
            "delta_pct":     delta_pct,
            "expiry_bucket": expiry_buck,
        })

    if not units:
        errors.append("No units parsed.")
        return {"errors": errors, "warnings": warnings}

    # ── Summary metrics ───────────────────────────────────────────────────
    total      = len(units)
    occupied   = sum(1 for u in units if u["status"] == "Occupied")
    vacant     = sum(1 for u in units if u["status"] == "Vacant")
    notice     = sum(1 for u in units if u["status"] == "Notice")
    model_adm  = sum(1 for u in units if u["status"] == "Model/Admin")
    phys_occ   = occupied / total if total > 0 else 0
    econ_occ_units = [u for u in units if u["effective_rent"] and u["effective_rent"] > 0]
    avg_inplace= np.mean([u["effective_rent"] for u in econ_occ_units]) if econ_occ_units else 0
    avg_market = np.mean([u["market_rent"]    for u in units if u["market_rent"]]) if units else 0
    loss_to_lease = avg_market - avg_inplace
    loss_to_lease_pct = (loss_to_lease / avg_market * 100) if avg_market > 0 else 0
    annual_sched = sum((u["effective_rent"] or 0) for u in units) * 12

    # ── Unit mix ──────────────────────────────────────────────────────────
    unit_mix: dict[str, dict] = {}
    for u in units:
        ut = u["unit_type"]
        if ut not in unit_mix:
            unit_mix[ut] = {"count": 0, "inplace_rents": [], "market_rents": [], "sizes": []}
        unit_mix[ut]["count"] += 1
        if u["effective_rent"] and u["effective_rent"] > 0:
            unit_mix[ut]["inplace_rents"].append(u["effective_rent"])
        if u["market_rent"]:
            unit_mix[ut]["market_rents"].append(u["market_rent"])
        if u["unit_size_sf"]:
            unit_mix[ut]["sizes"].append(u["unit_size_sf"])

    for ut, d in unit_mix.items():
        d["pct"]         = d["count"] / total * 100 if total > 0 else 0
        d["avg_inplace"] = np.mean(d["inplace_rents"]) if d["inplace_rents"] else 0
        d["avg_market"]  = np.mean(d["market_rents"])  if d["market_rents"]  else 0
        d["avg_sf"]      = np.mean(d["sizes"])         if d["sizes"]         else 0
        d["avg_rent_sf"] = (d["avg_inplace"] / d["avg_sf"]) if d["avg_sf"] > 0 else 0

    # ── Lease expirations ─────────────────────────────────────────────────
    expiry_by_month: dict[str, int] = {}
    for u in units:
        if u["status"] == "Occupied" and u["lease_end"]:
            key = u["lease_end"].strftime("%Y-%m")
            expiry_by_month[key] = expiry_by_month.get(key, 0) + 1

    # Sort
    expiry_by_month = dict(sorted(expiry_by_month.items()))

    # ── Expiry bucket totals ──────────────────────────────────────────────
    bucket_totals: dict[str, int] = {}
    for u in units:
        b = u["expiry_bucket"]
        bucket_totals[b] = bucket_totals.get(b, 0) + 1

    # ── As-of date: last lease_end or today ──────────────────────────────
    valid_ends = [u["lease_end"] for u in units if u["lease_end"]]
    as_of_date = max(valid_ends) if valid_ends else date.today()

    return {
        "as_of_date": as_of_date,
        "units": units,
        "summary": {
            "total_units":       total,
            "occupied_units":    occupied,
            "vacant_units":      vacant,
            "notice_units":      notice,
            "model_admin_units": model_adm,
            "physical_occ":      phys_occ,
            "avg_inplace_rent":  avg_inplace,
            "avg_market_rent":   avg_market,
            "loss_to_lease":     loss_to_lease,
            "loss_to_lease_pct": loss_to_lease_pct,
            "annual_sched_rent": annual_sched,
        },
        "unit_mix":           unit_mix,
        "lease_expirations":  expiry_by_month,
        "expiry_buckets":     bucket_totals,
        "errors":   errors,
        "warnings": warnings,
    }

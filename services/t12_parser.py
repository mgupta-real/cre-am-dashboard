"""
services/t12_parser.py
Parses T12 Excel files matching the CRE dashboard canonical format.

Expected structure (from Phoenix Commons T12 template):
 Row 4  → "T12 As Of Date:" in col C, date in col D
 Row 8  → header row: col C = "Category", col D = "T12 Line-Item Name",
           cols E-P = monthly dates (12 months), col R = T12, col S = T6,
           col T = T3, col U = T1
 Row 9+ → data rows
"""
import re
from datetime import datetime
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import load_workbook


# ── Revenue / Expense category classification ─────────────────────────────
REVENUE_KEYWORDS = {
    "gross potential", "vacancy loss", "loss to lease", "non-revenue",
    "bad debt", "concessions", "rubs", "fee income", "parking income",
    "late fee", "cable", "pet charge", "attorney", "renter", "misc",
    "laundry", "rental income", "net rental", "total revenue", "total income",
    "total effective", "other income"
}
EXPENSE_KEYWORDS = {
    "personnel", "payroll", "utilities", "repairs", "maintenance",
    "turnover", "contract services", "advertising", "administrative",
    "management fee", "real estate tax", "insurance", "landscaping",
    "marketing", "controllable", "non controllable", "operating expenses"
}
NOI_KEYWORDS = {"net operating income", "noi"}


def _is_revenue(cat: str, item: str) -> bool:
    text = f"{cat or ''} {item or ''}".lower()
    return any(k in text for k in REVENUE_KEYWORDS)


def _is_expense(cat: str, item: str) -> bool:
    text = f"{cat or ''} {item or ''}".lower()
    return any(k in text for k in EXPENSE_KEYWORDS)


def _is_noi(item: str) -> bool:
    return any(k in (item or "").lower() for k in NOI_KEYWORDS)


def _safe_float(v) -> float | None:
    if v is None or v == "" or v is False:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_t12(file_path: str) -> dict:
    """
    Parse a T12 Excel file and return a structured dict.

    Returns
    -------
    {
        "as_of_date": datetime,
        "month_dates": [datetime, ...],        # up to 12 months
        "line_items": [
            {
                "category": str,
                "line_item": str,
                "monthly": [float|None, ...],  # aligned to month_dates
                "t12": float|None,
                "t6": float|None,
                "t3": float|None,
                "t1": float|None,
                "confidence": float|None,
                "is_revenue": bool,
                "is_expense": bool,
                "is_noi": bool,
                "is_subtotal": bool,
            }, ...
        ],
        "summary": {
            "total_revenue_t12": float,
            "total_expenses_t12": float,
            "noi_t12": float,
            "noi_margin_t12": float,
            "total_revenue_t6": float,
            "total_expenses_t6": float,
            "noi_t6": float,
            "total_revenue_t3": float,
            "total_expenses_t3": float,
            "noi_t3": float,
            "total_revenue_t1": float,
            "total_expenses_t1": float,
            "noi_t1": float,
        },
        "monthly_totals": {
            "revenue": [float, ...],
            "expenses": [float, ...],
            "noi": [float, ...],
        },
        "revenue_mix": {category: float},   # T12
        "expense_mix": {category: float},   # T12
        "errors": [str],
        "warnings": [str],
    }
    """
    errors = []
    warnings = []

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        return {"errors": [f"Cannot open file: {e}"], "warnings": []}

    if "T12" not in wb.sheetnames:
        # Try first sheet
        ws = wb.active
        warnings.append("Sheet 'T12' not found — using first sheet.")
    else:
        ws = wb["T12"]

    all_rows = list(ws.iter_rows(values_only=True))

    # ── Step 1: Find as-of date ───────────────────────────────────────────
    as_of_date = None
    header_row_idx = None
    month_col_start = None  # 0-based column index of first monthly column
    t12_col = t6_col = t3_col = t1_col = None
    conf_col = None
    cat_col = None
    item_col = None

    for i, row in enumerate(all_rows):
        for j, cell in enumerate(row):
            if isinstance(cell, str) and "T12 As Of Date" in cell:
                # as_of_date is next non-None value in same row
                for k in range(j + 1, min(j + 5, len(row))):
                    if isinstance(row[k], datetime):
                        as_of_date = row[k]
                        break
            if isinstance(cell, str) and cell.strip() == "Category":
                cat_col = j
                # Look right for "T12 Line-Item Name"
                for k in range(j + 1, min(j + 5, len(row))):
                    if isinstance(row[k], str) and "line" in row[k].lower():
                        item_col = k
                        break
                # Monthly columns start after item_col
                month_cols = []
                for k in range((item_col or j) + 1, len(row)):
                    if isinstance(row[k], datetime):
                        month_cols.append((k, row[k]))
                if month_cols:
                    header_row_idx = i
                    month_col_start = month_cols[0][0]
                    month_dates = [d for _, d in month_cols[:12]]
                    month_col_indices = [c for c, _ in month_cols[:12]]
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        errors.append("Could not locate header row with 'Category' column.")
        return {"errors": errors, "warnings": warnings}

    if as_of_date is None:
        warnings.append("T12 as-of date not found; using last month date.")
        as_of_date = month_dates[-1] if month_dates else None

    # Find T12/T6/T3/T1 columns and Confidence Score
    hdr = all_rows[header_row_idx]
    for j, cell in enumerate(hdr):
        if isinstance(cell, str):
            cl = cell.strip().upper()
            if cl == "T12" and t12_col is None and j > month_col_indices[-1]:
                t12_col = j
            elif cl == "T6" and t6_col is None and j > month_col_indices[-1]:
                t6_col = j
            elif cl == "T3" and t3_col is None and j > month_col_indices[-1]:
                t3_col = j
            elif cl in ("T1", "CURRENT MTD") and t1_col is None and j > month_col_indices[-1]:
                t1_col = j
            elif "confidence" in cl.lower() and conf_col is None:
                conf_col = j

    # If not found by header, guess positions relative to month end
    if t12_col is None:
        t12_col = month_col_indices[-1] + 2  # skip 1 blank
    if t6_col is None:
        t6_col = t12_col + 1
    if t3_col is None:
        t3_col = t6_col + 1
    if t1_col is None:
        t1_col = t3_col + 1

    # ── Step 2: Parse data rows ───────────────────────────────────────────
    line_items = []
    prev_category = None

    for row in all_rows[header_row_idx + 1:]:
        # Skip completely empty rows
        non_null = [v for v in row if v is not None and v != ""]
        if not non_null:
            continue

        cat_val  = row[cat_col]  if cat_col  < len(row) else None
        item_val = row[item_col] if item_col < len(row) else None

        # Skip meta/helper rows (non-string item names, boolean False, etc.)
        if item_val is None or item_val is False or isinstance(item_val, (int, float, datetime)):
            continue
        item_str = str(item_val).strip()
        if not item_str or len(item_str) < 2:
            continue

        # Track last seen category
        if cat_val and isinstance(cat_val, str) and cat_val.strip():
            prev_category = cat_val.strip()
        cat_str = prev_category or ""

        # Monthly values
        monthly = []
        for ci in month_col_indices:
            monthly.append(_safe_float(row[ci]) if ci < len(row) else None)

        t12 = _safe_float(row[t12_col]) if t12_col < len(row) else None
        t6  = _safe_float(row[t6_col])  if t6_col  < len(row) else None
        t3  = _safe_float(row[t3_col])  if t3_col  < len(row) else None
        t1  = _safe_float(row[t1_col])  if t1_col  < len(row) else None
        conf = _safe_float(row[conf_col]) if conf_col and conf_col < len(row) else None

        # Subtotal rows have a blank category and repeated item name
        is_sub = (cat_val is None or str(cat_val).strip() == "")

        li = {
            "category":   cat_str,
            "line_item":  item_str,
            "monthly":    monthly,
            "t12":        t12,
            "t6":         t6,
            "t3":         t3,
            "t1":         t1,
            "confidence": conf,
            "is_revenue": _is_revenue(cat_str, item_str),
            "is_expense": _is_expense(cat_str, item_str),
            "is_noi":     _is_noi(item_str),
            "is_subtotal":is_sub,
        }
        line_items.append(li)

    if not line_items:
        errors.append("No line items could be parsed from this file.")
        return {"errors": errors, "warnings": warnings}

    # ── Step 3: Build summary ─────────────────────────────────────────────
    def find_value(keyword: str, col: str = "t12") -> float | None:
        """Search all line items (including subtotals) for the keyword."""
        kl = keyword.lower()
        for li in line_items:
            if kl in li["line_item"].lower():
                v = li.get(col)
                if v is not None:
                    return v
        return None

    total_rev_t12  = find_value("total revenue", "t12") or find_value("total income", "t12")
    total_exp_t12  = find_value("operating expenses", "t12")
    noi_t12        = find_value("net operating income", "t12")

    total_rev_t6   = find_value("total revenue", "t6") or find_value("total income", "t6")
    total_exp_t6   = find_value("operating expenses", "t6")
    noi_t6         = find_value("net operating income", "t6")

    total_rev_t3   = find_value("total revenue", "t3") or find_value("total income", "t3")
    total_exp_t3   = find_value("operating expenses", "t3")
    noi_t3         = find_value("net operating income", "t3")

    total_rev_t1   = find_value("total revenue", "t1") or find_value("total income", "t1")
    total_exp_t1   = find_value("operating expenses", "t1")
    noi_t1         = find_value("net operating income", "t1")

    noi_margin = (noi_t12 / total_rev_t12) if (noi_t12 and total_rev_t12 and total_rev_t12 != 0) else None

    # ── Step 4: Monthly totals ─────────────────────────────────────────────
    # Use Total Revenue / Operating Expenses rows directly — single source of
    # truth that avoids double-counting individual line items.
    n_months = len(month_dates)
    monthly_rev = [0.0] * n_months
    monthly_exp = [0.0] * n_months

    REV_TOTAL_NAMES = {"total revenue", "total income", "total effective income"}
    EXP_TOTAL_NAMES = {"operating expenses", "total operating expenses", "total expenses"}

    for li in line_items:
        if li["line_item"].lower().strip() in REV_TOTAL_NAMES:
            vals = [li["monthly"][m] if m < len(li["monthly"]) else None for m in range(n_months)]
            if any(v and v > 0 for v in vals):
                monthly_rev = [v if (v and v > 0) else 0.0 for v in vals]
                break

    for li in line_items:
        if li["line_item"].lower().strip() in EXP_TOTAL_NAMES:
            vals = [li["monthly"][m] if m < len(li["monthly"]) else None for m in range(n_months)]
            if any(v and v > 0 for v in vals):
                monthly_exp = [abs(v) if (v and v > 0) else 0.0 for v in vals]
                break

    # Fallback: sum non-subtotal items (skip top-level catch-alls to avoid double count)
    SKIP_NAMES = {"residential income", "rental income", "gross potential rents"}
    if sum(monthly_rev) == 0:
        for li in line_items:
            if li["is_revenue"] and not li["is_subtotal"] and li.get("t12") and li["t12"] > 0:
                if li["line_item"].lower().strip() in SKIP_NAMES:
                    continue
                for m in range(n_months):
                    v = li["monthly"][m] if m < len(li["monthly"]) else None
                    if v and v > 0:
                        monthly_rev[m] += v

    if sum(monthly_exp) == 0:
        for li in line_items:
            if li["is_expense"] and not li["is_subtotal"] and li.get("t12") and li["t12"] > 0:
                for m in range(n_months):
                    v = li["monthly"][m] if m < len(li["monthly"]) else None
                    if v and v > 0:
                        monthly_exp[m] += abs(v)

    monthly_noi = [monthly_rev[m] - monthly_exp[m] for m in range(n_months)]

    # ── Step 5: Revenue & expense mix (T12) ───────────────────────────────
    rev_mix: dict[str, float] = {}
    exp_mix: dict[str, float] = {}

    rev_cats = {
        "Gross Potential Rents": ["gross potential", "residential income"],
        "RUBS": ["rubs", "utility rebill"],
        "Fee Income": ["fee income", "administrative fees", "amenity fees", "application fees",
                       "storage rent", "pest control income"],
        "Parking": ["parking income", "garage"],
        "Late Fees": ["late fee", "nsf", "termination fee"],
        "Pet Charge": ["pet charge", "pet rent"],
        "Other Income": ["other income", "cable", "renter", "misc", "attorney"],
    }
    exp_cats_map = {
        "Personnel": ["personnel", "payroll"],
        "Utilities": ["utilities", "util "],
        "Repairs & Maintenance": ["repairs & maintenance", "turnover"],
        "Contract Services": ["contract services"],
        "Advertising & Marketing": ["advertising", "marketing"],
        "Administrative": ["administrative"],
        "Management Fees": ["management fee"],
        "Real Estate Taxes": ["real estate tax", "re tax"],
        "Insurance": ["insurance"],
        "Landscaping": ["landscaping"],
    }

    for li in line_items:
        if li["is_subtotal"] or li["t12"] is None:
            continue
        iname = li["line_item"].lower()
        cat   = li["category"].lower()
        val   = li["t12"]

        if val and val > 0 and li["is_revenue"]:
            placed = False
            for bucket, kws in rev_cats.items():
                if any(k in iname or k in cat for k in kws):
                    rev_mix[bucket] = rev_mix.get(bucket, 0) + val
                    placed = True
                    break
            if not placed and iname not in ("total revenue", "total income", "total effective income",
                                              "net rental income", "rental income"):
                rev_mix["Other Income"] = rev_mix.get("Other Income", 0) + val

        if val and val > 0 and li["is_expense"]:
            placed = False
            for bucket, kws in exp_cats_map.items():
                if any(k in iname or k in cat for k in kws):
                    exp_mix[bucket] = exp_mix.get(bucket, 0) + val
                    placed = True
                    break
            if not placed and iname not in ("operating expenses", "controllable", "non controllable"):
                exp_mix["Other"] = exp_mix.get("Other", 0) + val

    return {
        "as_of_date":    as_of_date,
        "month_dates":   month_dates,
        "line_items":    line_items,
        "summary": {
            "total_revenue_t12":  total_rev_t12,
            "total_expenses_t12": total_exp_t12,
            "noi_t12":            noi_t12,
            "noi_margin_t12":     noi_margin,
            "total_revenue_t6":   total_rev_t6,
            "total_expenses_t6":  total_exp_t6,
            "noi_t6":             noi_t6,
            "total_revenue_t3":   total_rev_t3,
            "total_expenses_t3":  total_exp_t3,
            "noi_t3":             noi_t3,
            "total_revenue_t1":   total_rev_t1,
            "total_expenses_t1":  total_exp_t1,
            "noi_t1":             noi_t1,
        },
        "monthly_totals": {
            "revenue":  monthly_rev,
            "expenses": monthly_exp,
            "noi":      monthly_noi,
        },
        "revenue_mix":  rev_mix,
        "expense_mix":  exp_mix,
        "errors":       errors,
        "warnings":     warnings,
    }
